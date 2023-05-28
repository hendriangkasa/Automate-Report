import pandas as pd #pandas untuk membuat dataframe(df)
from openpyxl import load_workbook #untuk berinterkasi antara python & excel file
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import json
import string


class ExcelReportPlugin():
    def __init__(self,
                 input_file,
                 output_file):
        self.input_file = input_file
        self.output_file = output_file
    
    def main(self):
        df = self.read_input_file()
        df_transform = self.transform(df)
        self.create_output_file(df_transform)
        print('workbook created')

        wb = load_workbook(self.output_file)
        wb.active = wb['Report']

        min_column = wb.active.min_column
        max_column = wb.active.max_column
        min_row = wb.active.min_row
        max_row = wb.active.max_row
        self.barchart(wb.active, min_column, max_column, min_row, max_row)
        self.add_total(wb.active, max_column. min_row, max_row)
        self.save_file(wb)

    def read_input_file(self):
        df = pd.read_excel(self.input_file)
        print(df.head())
        return df
    
    def transform(self, df):
        df_transform = df.pivot_table(index='Gambar',
                                      columns='Product line',
                                      values='Total',
                                      aggfunc='sum').round()
        print(df_transform)

    def create_output_file(self, df):
        print('Save dataframe to excel...')

        df.to_excel(self.output_file, 
                    sheet_name='Report', 
                    startrow=4)
        
        print('Save dataframe done...')

    def barchart(self, workbook, min_column, max_column, min_row, max_row):
        barchart = BarChart()

        data = Reference(workbook, 
                        min_col=min_column+1,
                        max_col=max_column,
                        min_row=min_row,
                        max_row=max_row
                        )

        categories = Reference(workbook,
                                min_col=min_column,
                                max_col=min_column,
                                min_row=min_row+1,
                                max_row=max_row
                                )

        barchart.add_data(data, titles_from_data=True)
        barchart.set_categories(categories)


        wb.active.add_chart(barchart, 'B12')
        barchart.title = 'Sales berdasarkan Produk'
        barchart.style = 2

    def add_total(self, workbook, max_column. min_row, max_row):
        alphabet = list(string.ascii_uppercase)
        alphabet_excel = alphabet[:max_column]
        #[A,B,C,D,E,F,G]
        for i in alphabet_excel:
            if i != 'A':
                workbook[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
                workbook[f'{i}{max_row+1}'].style = 'Currency'

        workbook[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

        workbook['A1'] = 'Sales Report'
        workbook['A2'] = '2019'
        workbook['A1'].font = Font('Arial', bold=True, size=20)
        workbook['A2'].font = Font('Arial', bold=True, size=10)

    def save_file(self, wb):
        wb.save(output_file)
        print('File saved')